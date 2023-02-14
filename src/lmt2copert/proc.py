"""Processing Module
"""

from logging import getLogger
from json import loads, dumps
from os.path import join
from math import ceil
from re import findall

import pandas as pd
from openpyxl import load_workbook


logger = getLogger("lmt2copert.proc")

def run_model(lmtjson, copert_xlsx, xml_in, vehicle, outdir):
    """Runs the model
    """
    df = pd.read_excel(copert_xlsx, "URBAN_OFF_PEAK_SPEED")
    urban_off_peak_speed = df.iloc[:,4]
    df2 = pd.read_excel(copert_xlsx, "URBAN_PEAK_SPEED")
    urban_peak_speed = df2.iloc[:,4]
    df3 = pd.read_excel(copert_xlsx, "URBAN_PEAK_SHARE")
    urban_peak_share = df3.iloc[:,4]
    df4 = pd.read_excel(copert_xlsx, "URBAN_OFF_PEAK_SHARE")
    urban_off_peak_share = df4.iloc[:,4]

    workbook = load_workbook(filename="lmt_LEAD_input_to_COPERT.xlsx")
    # open workbook
    sheet = workbook.active

    sheet_names = ["URBAN_OFF_PEAK_SPEED", "URBAN_PEAK_SPEED",
                   "URBAN_OFF_PEAK_SHARE", "URBAN_PEAK_SHARE",
                   "STOCK", "MEAN_ACTIVITY"]
    for sheet in sheet_names:
        workbook[sheet].delete_rows(2, 12)
    workbook.save(filename = copert_xlsx)

    num = 100
    if vehicle == 'diesel':
        category = 'Light Commercial Vehicles'
        fuel = 'Diesel'
        segment = 'N1-III'
        num = 161
        eurostandard = 'Euro 6 a/b/c'
    elif vehicle == 'hybrid':
        category = 'Passenger Cars'
        fuel = 'Petrol Hybrid'
        segment = 'Large-SUV-Executive'
        eurostandard = 'Euro 6 a/b/c'
        num = 161
    elif vehicle == 'nv200':
        category = 'HeavyCommercial Vehicles'
        fuel = 'Diesel'
        segment = 'N1-III'
        eurostandard = 'Euro 6 a/b/c'
        num = 100

    dd = pd.read_json(lmtjson)
    dd['Category'] = category
    dd['Fuel'] = fuel
    dd['Segment'] = segment
    dd['EuroStandard'] = eurostandard

    with open(xml_in, 'r') as f:
        a = f.read()
        xml_pattern = "(?:<NumberOfServices.*?>)(.*?)(?:<\\/NumberOfServices>)"
        val = findall(xml_pattern, a)[0]

    dd['Stock'] = ceil(int(val) / num)
    dd['MeanActivity'] = 60             # FIXME - 60 ?

    with open(join(outdir, "output.json"), "w") as ff:
        data = loads(dd.to_json(orient = "records"))
        ff.truncate()
        ff.write(dumps(data))

    d = pd.read_json(lmtjson)
    # FIXME - 2021 ?
    d = d.rename(columns = {"Stock":"2021", "MeanActivity":"2021_"})
    d2 = d[["Category", "Fuel", "Segment", "EuroStandard"]]
    d3 =  d[["Category", "Fuel", "Segment", "EuroStandard","2021"]]
    d4 = d[["Category", "Fuel", "Segment", "EuroStandard","2021_"]]
    d4 = d4.rename(columns = {"2021_":"2021"})
    if len(urban_peak_speed) != d2.shape[0]:
        if len(urban_peak_speed) > d2.shape[0]:
            urban_peak_speed = urban_peak_speed[:d2.shape[0]]
            urban_off_peak_speed = urban_off_peak_speed[:d2.shape[0]]
            urban_peak_share = urban_peak_share[:d2.shape[0]]
            urban_off_peak_share = urban_off_peak_share[:d2.shape[0]]
        else:
            while len(urban_peak_speed) != d2.shape[0]:
                urban_peak_speed =   urban_peak_speed.append(pd.Series(urban_peak_speed[0]), ignore_index = True)
                urban_off_peak_speed = urban_off_peak_speed.append(pd.Series(urban_off_peak_speed[0]), ignore_index = True)
                urban_peak_share = urban_peak_share.append(pd.Series(urban_peak_share[0]), ignore_index = True)
                urban_off_peak_share = urban_off_peak_share.append(pd.Series(urban_off_peak_share[0]), ignore_index = True)

    sheets = ["URBAN_OFF_PEAK_SPEED", "URBAN_PEAK_SPEED",
              "URBAN_OFF_PEAK_SHARE", "URBAN_PEAK_SHARE"]
    book = load_workbook(copert_xlsx)
    writer = pd.ExcelWriter(join(outdir, 'output.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    for sheetname in writer.sheets:
        if sheetname == "URBAN_OFF_PEAK_SPEED":
            d2[df.columns[4]] = urban_off_peak_speed.values
            d2.to_excel(writer,sheet_name=sheetname, startrow=0, index = False)
        elif sheetname == "URBAN_PEAK_SPEED":
            d2[df2.columns[4]] = urban_peak_speed.values
            d2.to_excel(writer,sheet_name=sheetname, startrow=0, index = False)
        elif sheetname == "URBAN_OFF_PEAK_SHARE":
            d2[df3.columns[4]] = urban_off_peak_share.values
            d2.to_excel(writer,sheet_name=sheetname, startrow=0, index = False)
        elif sheetname == "URBAN_PEAK_SHARE":
            d2[df4.columns[4]] = urban_peak_share.values
            d2.to_excel(writer,sheet_name=sheetname, startrow=0, index = False)
        elif sheetname == "STOCK":
            d3.to_excel(writer,sheet_name=sheetname, startrow=0, index = False)
        elif sheetname == "MEAN_ACTIVITY":
            d4.to_excel(writer,sheet_name=sheetname, startrow=0, index = False)
        else:
            pass

    writer.save()
